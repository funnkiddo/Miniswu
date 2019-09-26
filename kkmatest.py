from konlpy.tag import Kkma
from konlpy.utils import pprint
import re
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from konlpy.tag import Okt
from konlpy.tag import Kkma

kkma = Kkma()

import xlsxwriter

# workbook = xlsxwriter.Workbook("C:/python_source/section3/realresult.xlsx")
fworkbook = xlsxwriter.Workbook('C:/python_source/section3/nounsResult.xlsx')
# worksheet = workbook.add_worksheet()
fworksheet = fworkbook.add_worksheet()

# wbb = Workbook()
# sheet1 = wbb.active

okt = Okt()

wb = load_workbook('C:/python_source/section3/newsresult.xlsx')
sheet = wb['Sheet1']

row = 2
col = 1
frow = 0
fcol = 0

regex = re.compile(r'\d{3}-\d{3,4}-\d{4}')
# email = re.compile(r'[0-9a-zA-Z]([-_.]?[0-9a-zA-Z])*@[0-9a-zA-Z]([-_.]?[0-9a-zA-Z])*.[a-zA-Z]{2,3}')
personnum = re.compile(r'[0-9]{2}(0[1-9]|1[012])(0[1-9]|1[0-9]|2[0-9]|3[01])-?[012349][0-9]{6}')

sentences = list()

while True:
    text = sheet.cell(row=row, column=5).value

    if not text:
        break
        # row += 1

    # print(text)

    matchobj = re.search(regex, text)  # seaerch() : 처음 매칭되는 문자엶만 리턴, 모든 경우는 findall()이용
    matchobj3 = re.search(personnum, text)

    # print(matchobj)
    # print(matchobj3)
    # #
    sentences.append(text)
    #
    # phonenumber = matchobj.group()  # group() 실제 문제열 얻기위함
    # personnumsource = matchobj3.groups()
    #
    #
    #
    ins_cnt = 2


    if matchobj:
        poNum = "[" + "(" + "'" + format(matchobj.group()) + "'" + "]" + "," + " " + "'" + "Phone Num" + "'" + ")" + "]"
        print(poNum)
        # sheet1.append(matchobj.group())

        # sheet1.cell(row=row, column=1).value = matchobj
        fworksheet.write(frow, fcol, poNum)
    elif matchobj3:
        perNum = "[" + "(" + "'" + format(matchobj3.group()) + "'" + "]" + "," + " " + "'" + "Person Num" + "'" + ")" + "]"
        print(perNum)
        # sheet1.append(matchobj3.group())
        # sheet1.cell(row=row,column=1).value = matchobj3
        fworksheet.write(frow, fcol, perNum)
    else:
        poslist = kkma.pos(text)
       # if type == "NNG" or type == "NNP":


        print(format(poslist))
        # sheet1.append(poslist)
        # sheet1.cell(row=row,column=1).value = poslist
        fworksheet.write(frow, fcol, format(poslist))

    row += 1
    frow += 1


fworkbook.close()
# workbook.close()
# wbb.save(filename=file_name)
